import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename) # Open the given filename and return the workbook
    sheet = wb['Sheet1'] # specifying the sheet

    for row in range(2, sheet.max_row + 1): # sheet.max_row is a prop which returns how many rows are present there.
        cell = sheet.cell(row, 3) # specify any particular cell, .cell(row, col)
        corrected_price = cell.value * .9
        corrected_price_cell = sheet.cell(row, 4) # select a new cell, according to our need
        corrected_price_cell.value = corrected_price # set this set value

    # This Reference() funct will take sheet, and the specified row and col as args
    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart() # create a instance of BarChart() class
    chart.add_data(values) # values we take for draw our chart
    sheet.add_chart(chart, 'e2') # add the chart to the sheet, 'e2' will be the left-top corner of the chart

    wb.save(filename) # overriding the original file

process_workbook('transactions.xlsx')